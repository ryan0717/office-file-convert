import os
import re
import datetime
from markitdown import MarkItDown


# ============ Custom XLSX conversion (merged cells + number formats) / XLSX 自定义转换 ============

def _format_cell_value(cell):
    """Format cell value, preserving Excel number formats (percentage, currency, date, etc.).
    格式化单元格值，保留 Excel 中的数字格式（百分比、货币、日期等）。"""
    if cell.value is None:
        return ""

    val = cell.value
    fmt = cell.number_format

    # String / 字符串
    if isinstance(val, str):
        return val

    # Boolean / 布尔值
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"

    # Datetime (openpyxl auto-parses Excel date serial numbers to datetime objects)
    # 日期时间 (openpyxl auto-parses / 已自动解析)
    if isinstance(val, datetime.datetime):
        if val.hour or val.minute or val.second:
            return val.strftime("%Y-%m-%d %H:%M:%S")
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")

    # Numeric types / 数字类型
    if isinstance(val, (int, float)):
        # Percentage format: Excel stores as decimal (0.125 = 12.5%)
        if "%" in fmt:
            decimal_part = fmt.split(".")[1].replace("%", "") if "." in fmt else ""
            decimal_places = len(decimal_part) if decimal_part else 0
            return f"{val * 100:.{decimal_places}f}%"

        # Currency format / 货币格式
        for symbol in ("¥", "$", "€", "£"):
            if symbol in fmt:
                decimal_places = 0
                if "." in fmt:
                    after_dot = fmt.split(".")[1]
                    decimal_places = len(after_dot.replace("#", "0").replace(",", ""))
                formatted = f"{val:,.{decimal_places}f}"
                if "0" in fmt and fmt.find(symbol) < fmt.find("0"):
                    return f"{symbol}{formatted}"
                return f"{formatted}{symbol}"

        # Number format with thousands separator or decimal places / 数值格式（千分位或小数位）
        if "," in fmt or "." in fmt:
            decimal_places = 0
            if "." in fmt:
                after_dot = fmt.split(".")[1]
                decimal_places = len(after_dot.replace("#", "0").replace(",", ""))
            return f"{val:,.{decimal_places}f}"

        # Avoid .0 for integers with General format / General 格式的整数避免 .0
        if isinstance(val, float) and val == int(val) and fmt == "General":
            return str(int(val))

        return str(val)

    return str(val)


def _truncate_cell(text, max_len=80):
    """Truncate long cell text to keep markdown tables readable.
    截断过长的单元格文本，避免 markdown 表格列过宽不可读。"""
    text = text.replace("\n", " ")
    if len(text) <= max_len:
        return text
    return text[:max_len - 1] + "…"


def xlsx_to_markdown(file_path):
    """Convert XLSX file to Markdown with merged cell and number format support.
    将 XLSX 文件转为 Markdown，支持合并单元格和数字格式。

    Compared to MarkItDown's built-in XlsxConverter (pandas -> HTML -> markdownify):
    - Preserves merged cell content (repeats top-left value in all merged positions)
    - Preserves number formats (percentage, currency, thousands, dates, etc.)
    - Generates Markdown tables directly, avoiding HTML intermediate info loss
    - Handles full-row merges smartly (output once, avoid duplication)
    - Filters trailing empty columns to trim table width
    - Truncates long text for markdown readability
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    md_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_parts.append(f"## {sheet_name}\n")

        if ws.max_row is None or ws.max_column is None or ws.max_row == 0:
            continue

        # ---- Build merged cell mapping / 构建合并单元格映射 ----
        # Track each merged range and which rows are "full-row merges"
        merged_values = {}
        row_full_merge = {}  # row_idx -> True if row is covered by a full-width merge / 横跨所有列的合并区域

        for merged_range in ws.merged_cells.ranges:
            top_left_value = _format_cell_value(
                ws.cell(merged_range.min_row, merged_range.min_col)
            )
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = top_left_value
            # Detect full-row merge (spans col 1 to max_column) / 检测整行合并
            if (merged_range.min_col == 1
                    and merged_range.max_col >= ws.max_column):
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    row_full_merge[row] = top_left_value

        # ---- Build table data row by row / 逐行构建表格数据 ----
        rows = []
        for row_idx in range(1, ws.max_row + 1):
            # Full-row merge: output once as standalone paragraph / 整行合并：只输出一次
            if row_idx in row_full_merge:
                rows.append(("FULL_MERGE", row_full_merge[row_idx]))
                continue

            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) in merged_values:
                    row_data.append(str(merged_values[(row_idx, col_idx)]))
                else:
                    row_data.append(_format_cell_value(ws.cell(row_idx, col_idx)))
            # Skip fully empty rows / 跳过全空行
            if any(cell.strip() for cell in row_data):
                rows.append(("ROW", row_data))

        if not rows:
            continue

        # ---- Filter trailing empty columns / 过滤右侧空列 ----
        # Find max effective column count across all table rows
        table_rows = [r for r in rows if r[0] == "ROW"]
        if table_rows:
            max_data_col = 0
            for _, row_data in table_rows:
                # Scan right-to-left for last non-empty column / 从右向左找最后一个非空列
                for i in range(len(row_data) - 1, -1, -1):
                    if row_data[i].strip():
                        max_data_col = max(max_data_col, i + 1)
                        break
        else:
            max_data_col = 0

        # ---- Generate Markdown / 生成 Markdown ----
        table_started = False

        for item in rows:
            if item[0] == "FULL_MERGE":
                # Full-row merge: close current table, output as paragraph
                if table_started:
                    md_parts.append("")  # blank line after table / 表格后空行
                    table_started = False
                md_parts.append(item[1])
                md_parts.append("")
            else:
                _, row_data = item
                # Trim to effective column count / 裁剪到有效列数
                row_data = row_data[:max_data_col]
                # Pad to equal column count / 补齐列数
                row_data = row_data + [""] * (max_data_col - len(row_data))

                # Calculate column widths with truncation / 计算列宽（带截断）
                if not table_started:
                    # First table row: output header and separator / 首次遇到表格行
                    col_widths = [3] * max_data_col
                    for _, tr in table_rows:
                        tr_cropped = (tr + [""] * max_data_col)[:max_data_col]
                        for i, cell in enumerate(tr_cropped):
                            display = _truncate_cell(cell)
                            col_widths[i] = max(col_widths[i], len(display))

                    header = (
                        "| "
                        + " | ".join(
                            _truncate_cell(cell).ljust(col_widths[i])
                            for i, cell in enumerate(row_data)
                        )
                        + " |"
                    )
                    separator = "|" + "|".join("-" * (w + 2) for w in col_widths) + "|"
                    md_parts.append(header)
                    md_parts.append(separator)
                    table_started = True
                else:
                    line = (
                        "| "
                        + " | ".join(
                            _truncate_cell(cell).ljust(col_widths[i])
                            for i, cell in enumerate(row_data)
                        )
                        + " |"
                    )
                    md_parts.append(line)

        if table_started:
            md_parts.append("")  # blank line after table / 表格后空行

    wb.close()
    return "\n".join(md_parts).strip()


# ============ Legacy format conversion (.doc/.xls/.ppt -> modern) / 旧格式转换 ============

# Legacy -> modern format mapping / 旧格式 → 新格式的映射
OLD_FORMAT_MAP = {
    ".doc": ".docx",
    ".xls": ".xlsx",
    ".ppt": ".pptx",
}

# COM component FileFormat constants / COM 组件对应的 FileFormat 常量
COM_FILE_FORMATS = {
    ".doc": ("Word.Application", 16),    # wdFormatXMLDocument
    ".xls": ("Excel.Application", 51),   # xlOpenXMLWorkbook
    ".ppt": ("PowerPoint.Application", 24),  # ppSaveAsOpenXMLPresentation
}


def convert_old_format(src_path, src_ext, upgraded_dir, relative_path):
    """Convert legacy format file to modern format, save to upgraded dir, return new path.
    将旧格式文件转为新格式，保存到 upgraded 目录，返回新文件路径。

    Args:
        src_path: Full path of legacy format source file / 旧格式源文件完整路径
        src_ext: Source file extension (e.g. .doc) / 源文件扩展名
        upgraded_dir: Root path of upgraded folder / upgraded 文件夹根路径
        relative_path: Sub-path relative to input (e.g. 'subdir') / 相对于 input 的子路径

    Returns:
        Full path of converted file, or None on failure / 转换后的新格式文件完整路径，失败返回 None
    """
    import win32com.client

    new_ext = OLD_FORMAT_MAP[src_ext]
    app_name, file_format = COM_FILE_FORMATS[src_ext]

    # Build output path preserving directory structure / 构建输出路径（保持目录结构）
    target_dir = os.path.join(upgraded_dir, relative_path)
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)

    base_name = os.path.splitext(os.path.basename(src_path))[0]
    new_path = os.path.join(target_dir, base_name + new_ext)

    # Skip if modern format file already exists / 如果已存在同名新格式文件，跳过
    if os.path.exists(new_path):
        return new_path

    app = None
    try:
        app = win32com.client.Dispatch(app_name)
        # PowerPoint and Word/Excel have Visible property at different locations
        # PowerPoint 和 Word/Excel 的 Visible 属性位置不同
        if hasattr(app, "Visible"):
            app.Visible = False

        if src_ext == ".doc":
            doc = app.Documents.Open(src_path)
            doc.SaveAs2(new_path, FileFormat=file_format)
            doc.Close()
        elif src_ext == ".xls":
            wb = app.Workbooks.Open(src_path)
            wb.SaveAs(new_path, FileFormat=file_format)
            wb.Close()
        elif src_ext == ".ppt":
            pres = app.Presentations.Open(src_path)
            pres.SaveAs(new_path, FileFormat=file_format)
            pres.Close()

        return new_path

    except Exception as e:
        print(f"  [Legacy conversion failed / 旧格式转换失败] {os.path.basename(src_path)}: {e}")
        return None
    finally:
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass


# ============ Custom PDF conversion / PDF 自定义转换 ============

def pdf_to_markdown(file_path):
    """Convert PDF to Markdown using pymupdf4llm (preferred) or MarkItDown (fallback).
    将 PDF 转为 Markdown，优先使用 pymupdf4llm，失败时回退到 MarkItDown。

    pymupdf4llm advantages / 优势:
    - Better table detection and Markdown table generation / 更好的表格识别和 Markdown 表格生成
    - Preserves heading hierarchy with proper # syntax / 保留标题层级（# / ## / ###）
    - Handles Chinese text and fonts better / 更好地处理中文文本和字体
    - Extracts images as embedded base64 or file references / 提取图片为内嵌 base64 或文件引用
    """
    # Try pymupdf4llm first / 优先尝试 pymupdf4llm
    try:
        import pymupdf4llm
        md_text = pymupdf4llm.to_markdown(file_path)
        if md_text and md_text.strip():
            return md_text, os.path.splitext(os.path.basename(file_path))[0]
    except ImportError:
        print("  [Warning/警告] pymupdf4llm not installed, falling back to MarkItDown")
        print("  Run: pip install pymupdf4llm")
    except Exception as e:
        print(f"  [Warning/警告] pymupdf4llm failed ({e}), falling back to MarkItDown")

    # Fallback to MarkItDown / 回退到 MarkItDown
    from markitdown import MarkItDown
    md = MarkItDown()
    result = md.convert(file_path)
    return result.text_content, result.title or os.path.splitext(os.path.basename(file_path))[0]


def post_process_pdf(markdown):
    """Post-process PDF-specific conversion artifacts.
    后处理 PDF 转换产生的特殊问题。

    Fixes / 修复内容:
    - Remove (cid:x) placeholder tokens / 移除 (cid:x) 占位符
    - Clean up broken pipe characters in tables / 清理表格中散乱的 | 符号
    - Normalize heading formats / 标准化标题格式
    - Fix orphaned table fragments / 修复孤立的表格碎片
    """
    # Remove (cid:x) tokens (common in some PDF encodings) / 移除 (cid:x) 占位符
    markdown = re.sub(r"\(cid:\d+\)", "", markdown)

    # Remove standalone page numbers (e.g., lone "3" on a line) / 移除孤立的页码
    markdown = re.sub(r"^\d{1,3}$", "", markdown, flags=re.MULTILINE)

    # Clean up excessive pipe characters that aren't valid tables / 清理非表格的散乱竖线
    lines = markdown.split("\n")
    cleaned_lines = []
    for line in lines:
        # If line has pipes but doesn't look like a valid table row, clean it
        if "|" in line and not re.match(r"^\s*\|.*\|\s*$", line):
            # Check if it's a separator line (---|---|---)
            if not re.match(r"^\s*\|?[\s\-\|:]+\|?\s*$", line):
                # Replace standalone pipes that aren't part of table structure
                line = re.sub(r"(?<!\S)\|(?![-\s])", "", line)
        cleaned_lines.append(line)
    markdown = "\n".join(cleaned_lines)

    # Fix broken table rows: lines starting with | but missing closing | / 修复不完整的表格行
    lines = markdown.split("\n")
    fixed_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and not stripped.endswith("|") and "|" in stripped[1:]:
            # Likely a broken table row, add closing pipe
            stripped = stripped.rstrip() + " |"
        fixed_lines.append(stripped if stripped != line.strip() else line)
    markdown = "\n".join(fixed_lines)

    # Detect and convert number+text patterns to headings / 检测 "数字+文字" 模式转为标题
    # e.g., "四、监管技术手段" -> "## 四、监管技术手段"
    # e.g., "4.1 金税四期" -> "### 4.1 金税四期"
    lines = markdown.split("\n")
    heading_lines = []
    for line in lines:
        stripped = line.strip()
        # Match Chinese numbered headings: 一、 二、 三、 etc. / 匹配中文编号标题
        if re.match(r"^[一二三四五六七八九十]+、", stripped) and len(stripped) < 80:
            if not stripped.startswith("#"):
                stripped = "## " + stripped
        # Match decimal numbered headings: 4.1, 3.3, etc. / 匹配小数点编号标题
        elif re.match(r"^\d+\.\d+\s+\S", stripped) and len(stripped) < 80:
            if not stripped.startswith("#"):
                stripped = "### " + stripped
        heading_lines.append(stripped if stripped != line.strip() else line)
    markdown = "\n".join(heading_lines)

    # Remove excessive underline patterns (____) / 移除过多下划线装饰
    markdown = re.sub(r"_{5,}", "---", markdown)

    return markdown


# ============ Post-processing functions / 后处理函数 ============

def add_front_matter(markdown, title, source_file, file_type):
    """Add YAML front matter metadata to Markdown content.
    在 Markdown 内容前添加 YAML front matter 元数据。"""
    safe_title = title.replace('"', '\\"')
    safe_source = os.path.basename(source_file).replace('"', '\\"')
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    front_matter = (
        f"---\n"
        f'title: "{safe_title}"\n'
        f'source: "{safe_source}"\n'
        f'type: "{file_type}"\n'
        f'converted: "{now}"\n'
        f"---\n\n"
    )
    return front_matter + markdown


def post_process(markdown):
    """Post-process: normalize line endings + clean up blank lines.
    后处理：行尾标准化 + 空行清理。"""
    # Normalize line endings to LF / 统一换行符为 LF
    markdown = markdown.replace("\r\n", "\n").replace("\r", "\n")
    # Strip trailing whitespace from each line / 去除每行尾部空格
    markdown = "\n".join(line.rstrip() for line in markdown.split("\n"))
    # Collapse 3+ consecutive blank lines into 2 / 3 个及以上连续空行合并为 2 个换行
    markdown = re.sub(r"\n{3,}", "\n\n", markdown)
    # Strip leading/trailing whitespace / 去除首尾空白
    markdown = markdown.strip()
    return markdown


def fix_slide_numbering(markdown):
    """Convert PPTX HTML comment slide numbers to Markdown headings.
    将 PPTX 的 HTML 注释幻灯片编号改为 Markdown 标题。

    MarkItDown built-in output: <!-- Slide number: 1 -->
    Improved output:            ## 幻灯片 1
    """
    return re.sub(
        r"<!-- Slide number: (\d+) -->", r"## 幻灯片 \1", markdown
    )


# ============ DOCX style mapping / DOCX 样式映射 ============

# Cover both CN and EN heading style names so custom Word styles also map to Markdown heading levels
DOCX_STYLE_MAP = "\n".join([
    "p[style-name='Heading 1'] => #",
    "p[style-name='Heading 2'] => ##",
    "p[style-name='Heading 3'] => ###",
    "p[style-name='Heading 4'] => ####",
    "p[style-name='Heading 5'] => #####",
    "p[style-name='Heading 6'] => ######",
    "p[style-name='标题 1'] => #",
    "p[style-name='标题 2'] => ##",
    "p[style-name='标题 3'] => ###",
    "p[style-name='标题 4'] => ####",
    "p[style-name='标题 5'] => #####",
    "p[style-name='标题 6'] => ######",
])


# ============ Main conversion logic / 主转换逻辑 ============

def batch_convert():
    # ================= Configuration / 配置区域 =================
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Directory structure / 目录结构：input → output / upgraded
    input_dir = os.path.join(base_dir, "input")
    output_dir = os.path.join(base_dir, "output")
    upgraded_dir = os.path.join(base_dir, "upgraded")

    # Modern formats for direct conversion / 支持直接转换的新格式
    new_extensions = ('.pptx', '.docx', '.xlsx', '.pdf')
    # Legacy formats requiring upgrade / 需要先升级的旧格式
    old_extensions = tuple(OLD_FORMAT_MAP.keys())
    # All supported formats / 所有可处理的格式
    valid_extensions = new_extensions + old_extensions
    # ===========================================

    # ---- Startup banner / 启动横幅 ----
    print()
    print("=" * 50)
    print("  Office Files Batch to Markdown")
    print("  Office 文件批量转 Markdown 工具")
    print("=" * 50)
    print()
    print(f"  Working dir / 工作目录: {base_dir}")
    print(f"  Input dir / 输入目录  : {input_dir}")
    print(f"  Output dir / 输出目录 : {output_dir}")
    print()
    print("  Supported formats / 支持格式:")
    print("    Modern / 新格式: .pptx  .docx  .xlsx  .pdf")
    print("    Legacy / 旧格式: .doc   .xls   .ppt  (auto-upgrade / 自动升级)")
    print()

    # ---- Ensure directories exist / 确保目录存在 ----
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(upgraded_dir, exist_ok=True)

    # ---- Scan input files / 扫描输入文件 ----
    all_input_files = []
    for dirpath, dirnames, filenames in os.walk(input_dir):
        for filename in filenames:
            all_input_files.append(filename)

    if not all_input_files:
        print("[Info/提示] Input folder is empty / input 文件夹为空，没有需要转换的文件。")
        print()
        print("  Place Office files into 'input' and re-run.")
        print("  请将 Office 文件放入 input 文件夹后重新运行。")
        input("\nPress Enter to exit / 按回车键退出...")
        return

    # Categorize and display pending files / 分类展示待处理文件
    old_files = [f for f in all_input_files
                 if os.path.splitext(f)[1].lower() in old_extensions]
    new_files = [f for f in all_input_files
                 if os.path.splitext(f)[1].lower() in new_extensions]
    skip_files = [f for f in all_input_files
                  if os.path.splitext(f)[1].lower() not in valid_extensions]

    print(f"  Found / 扫描到 {len(all_input_files)} file(s) / 个文件：")
    if new_files:
        print(f"    Direct / 直接转换 : {len(new_files)}  {', '.join(new_files[:5])}"
              + ("  ..." if len(new_files) > 5 else ""))
    if old_files:
        print(f"    Upgrade / 需升级  : {len(old_files)}  {', '.join(old_files[:5])}"
              + ("  ..." if len(old_files) > 5 else ""))
    if skip_files:
        print(f"    Unsupported / 跳过: {len(skip_files)}  {', '.join(skip_files[:5])}"
              + ("  ..." if len(skip_files) > 5 else ""))
    print()

    md = MarkItDown()
    log_data = {
        "success": [],
        "upgraded": [],
        "skipped": [],
        "failed": []
    }

    # ========== Phase 1: Legacy format upgrade / 阶段一：旧格式升级 ==========
    old_files_found = False
    for dirpath, dirnames, filenames in os.walk(input_dir):
        for filename in filenames:
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext in old_extensions:
                if not old_files_found:
                    print()
                    print("=" * 50)
                    print("  Phase 1: Legacy Format Upgrade (.doc/.xls/.ppt)")
                    print("  阶段一：旧格式文件升级")
                    print("  Note: Requires Microsoft Office, no popup dialogs")
                    print("  提示：需要 Office 程序且无弹窗阻塞")
                    print("=" * 50)
                    old_files_found = True

                src_file_path = os.path.join(dirpath, filename)
                relative_path = os.path.relpath(dirpath, input_dir)
                if relative_path == ".":
                    relative_path = ""

                print(f"  Upgrade/升级: {filename} → {OLD_FORMAT_MAP[file_ext]}")
                new_path = convert_old_format(
                    src_file_path, file_ext, upgraded_dir, relative_path
                )
                if new_path:
                    log_data["upgraded"].append(
                        f"{filename} → {os.path.basename(new_path)}"
                    )
                else:
                    log_data["failed"].append(
                        f"{filename} (upgrade failed / 旧格式升级失败)"
                    )

    if old_files_found:
        if log_data["upgraded"]:
            print(f"\n  Upgrade done / 旧格式升级完成: {len(log_data['upgraded'])} succeeded")
        if any("upgrade failed" in f for f in log_data["failed"]):
            fail_old = [f for f in log_data["failed"] if "upgrade failed" in f]
            print(f"  Upgrade failed / 旧格式升级失败: {len(fail_old)} file(s)")
    else:
        print("  No legacy files found, skipping Phase 1 / 无旧格式文件，跳过阶段一")

    # ========== Phase 2: Convert to Markdown / 阶段二：转 Markdown ==========
    print()
    print("=" * 50)
    print("  Phase 2: Convert to Markdown / 阶段二：文件转 Markdown")
    print("=" * 50)

    # Collect files to convert / 收集需要转换的文件列表
    files_to_convert = []

    # Modern format files: read directly from input / 新格式文件：直接从 input 读取
    for dirpath, dirnames, filenames in os.walk(input_dir):
        for filename in filenames:
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext in new_extensions:
                src_file_path = os.path.join(dirpath, filename)
                relative_path = os.path.relpath(dirpath, input_dir)
                if relative_path == ".":
                    relative_path = ""
                files_to_convert.append((src_file_path, relative_path, filename, file_ext))

    # Legacy upgraded files: read from upgraded / 旧格式升级后的文件：从 upgraded 读取
    if os.path.exists(upgraded_dir):
        for dirpath, dirnames, filenames in os.walk(upgraded_dir):
            for filename in filenames:
                file_ext = os.path.splitext(filename)[1].lower()
                if file_ext in new_extensions:
                    src_file_path = os.path.join(dirpath, filename)
                    relative_path = os.path.relpath(dirpath, upgraded_dir)
                    if relative_path == ".":
                        relative_path = ""
                    files_to_convert.append((src_file_path, relative_path, filename, file_ext))

    if not files_to_convert:
        print("  No convertible files found / 没有可转换的文件。")
    else:
        print(f"  {len(files_to_convert)} file(s) to convert / 个文件待转换，开始处理...\n")

    for src_file_path, relative_path, filename, file_ext in files_to_convert:
        target_dir = os.path.join(output_dir, relative_path)

        try:
            output_filename = os.path.splitext(filename)[0] + ".md"
            output_full_path = os.path.join(target_dir, output_filename)

            if not os.path.exists(target_dir):
                os.makedirs(target_dir)

            file_type = file_ext.lstrip(".")

            # --- Select conversion strategy by file type / 根据文件类型选择转换策略 ---
            if file_ext == ".xlsx":
                # XLSX: Custom converter (merged cells + number formats) / 自定义转换器
                markdown = xlsx_to_markdown(src_file_path)
                title = os.path.splitext(filename)[0]

            elif file_ext == ".docx":
                # DOCX: MarkItDown + custom style_map (CN/EN heading mapping) / 中英文标题映射
                result = md.convert(
                    src_file_path, style_map=DOCX_STYLE_MAP
                )
                markdown = result.text_content
                title = result.title or os.path.splitext(filename)[0]

            elif file_ext == ".pptx":
                # PPTX: MarkItDown + fix slide numbering / 修复幻灯片编号
                result = md.convert(src_file_path)
                markdown = result.text_content
                markdown = fix_slide_numbering(markdown)
                title = result.title or os.path.splitext(filename)[0]

            elif file_ext == ".pdf":
                # PDF: Custom converter (pymupdf4llm + post-processing) / PDF 专用转换器
                markdown, title = pdf_to_markdown(src_file_path)
                markdown = post_process_pdf(markdown)
                if not title:
                    title = os.path.splitext(filename)[0]

            else:
                # Other formats: MarkItDown default / 其他格式默认转换
                result = md.convert(src_file_path)
                markdown = result.text_content
                title = result.title or os.path.splitext(filename)[0]

            # --- Post-processing / 后处理 ---
            markdown = post_process(markdown)

            # --- Add YAML front matter / 添加 YAML front matter ---
            markdown = add_front_matter(
                markdown, title, src_file_path, file_type
            )

            with open(output_full_path, "w", encoding="utf-8") as f:
                f.write(markdown)

            log_data["success"].append(filename)
            print(f"  ✓ {filename}")

        except Exception as e:
            error_msg = f"{filename} (error/错误: {str(e)})"
            log_data["failed"].append(error_msg)
            print(f"  ✗ {filename}: {e}")

    # Count skipped files / 统计跳过的文件
    for dirpath, dirnames, filenames in os.walk(input_dir):
        for filename in filenames:
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext not in valid_extensions:
                log_data["skipped"].append(filename)

    # ================= Generate log file / 生成日志文件 =================
    log_filename = (
        "convert_log_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".txt"
    )
    log_file_path = os.path.join(base_dir, log_filename)

    with open(log_file_path, "w", encoding="utf-8") as f:
        f.write("=" * 50 + "\n")
        f.write(
            f"Conversion complete / 任务完成 - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        )
        f.write("=" * 50 + "\n\n")

        f.write(f"Markdown converted / 转换成功: {len(log_data['success'])}\n")
        f.write(f"Legacy upgraded / 旧格式升级: {len(log_data['upgraded'])}\n")
        f.write(f"Skipped / 格式跳过          : {len(log_data['skipped'])}\n")
        f.write(f"Failed / 转换失败            : {len(log_data['failed'])}\n\n")

        if log_data["upgraded"]:
            f.write("-" * 50 + "\n")
            f.write("Upgrade details / 旧格式升级详情:\n")
            for item in log_data["upgraded"]:
                f.write(f"  {item}\n")
            f.write("\n")

        if log_data["failed"]:
            f.write("-" * 50 + "\n")
            f.write("Failed files / 转换失败的文件:\n")
            for item in log_data["failed"]:
                f.write(f"  {item}\n")
            f.write("\n")

        if log_data["skipped"]:
            f.write("-" * 50 + "\n")
            f.write("Skipped files / 被跳过的文件:\n")
            for item in log_data["skipped"]:
                f.write(f"  {item}\n")

    # ---- Summary / 完成汇总 ----
    print()
    print("=" * 50)
    print("  All done! / 全部任务完成！")
    print("=" * 50)
    print()
    print(f"  Converted / 转换成功: {len(log_data['success'])}")
    print(f"  Upgraded / 格式升级 : {len(log_data['upgraded'])}")
    print(f"  Skipped / 格式跳过  : {len(log_data['skipped'])}")
    print(f"  Failed / 转换失败   : {len(log_data['failed'])}")
    print()
    print(f"  Output dir / 输出目录: {output_dir}")
    print(f"  Log file / 日志文件  : {log_file_path}")
    if log_data["success"]:
        print()
        print("  Generated Markdown files / 已生成的 Markdown 文件：")
        for name in log_data["success"]:
            print(f"    - {os.path.splitext(name)[0]}.md")
    if log_data["failed"]:
        print()
        print("  Failed files, please check / 以下文件转换失败，请检查：")
        for item in log_data["failed"]:
            print(f"    - {item}")
    print()
    input("Press Enter to exit / 按回车键退出...")


if __name__ == "__main__":
    batch_convert()
